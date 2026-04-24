"""
Banyan Throwdown — Master Tally Sheet (A3 LANDSCAPE).
For the owner to track every athlete's score as cards come in.
Two tables side-by-side so every row is tall enough to write in.
Optimized for 7-Eleven Japan's A3 color print (¥100/sheet).
"""
import glob
import os
import openpyxl
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import mm, inch
from reportlab.lib.colors import HexColor, black, white
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "BanyanThrowdown_TallySheet_A3.pdf")
LOGO = os.path.join(HERE, "logo.png")

pdfmetrics.registerFont(
    TTFont("IPAGothic", "/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf")
)
JP = "IPAGothic"

# Palette — matches the scorecards
INK      = HexColor("#0E0E0E")
MUTED    = HexColor("#555555")
BANYAN   = HexColor("#2E6B34")
BANYAN_D = HexColor("#1F4D2B")
BANYAN_L = HexColor("#E6F0E5")
ACCENT   = HexColor("#E07A1F")
PAGE_BG  = HexColor("#ECE8DD")
ROW_ALT  = HexColor("#F5F2E8")
RULE     = HexColor("#8A8272")
GRID     = HexColor("#D0C9B8")

PAGE_W, PAGE_H = landscape(A3)   # 420 x 297 mm
OUTER = 0.18 * inch
M = 0.32 * inch

# ------------------------------------------------------------
# Roster
# ------------------------------------------------------------
def load_roster():
    xlsx_candidates = glob.glob(os.path.join(HERE, "*.xlsx"))
    if not xlsx_candidates:
        raise FileNotFoundError("no xlsx in " + HERE)
    wb = openpyxl.load_workbook(xlsx_candidates[0], data_only=True)
    ws = wb["Schedule"]
    roster = []
    for row in range(6, 15):
        div = ws.cell(row=row, column=3).value
        heat = ws.cell(row=row, column=4).value
        heat_num = int(heat.split()[1])
        for lane in range(1, 7):
            name = ws.cell(row=row, column=4 + lane).value
            if name:
                roster.append((div, heat_num, lane, name.strip()))
    return roster

roster = load_roster()
order = ["Scaled (F)", "Scaled (M)", "RX (F)", "RX (M)"]
by_div = {d: [] for d in order}
for d, h, l, n in roster:
    by_div[d].append((h, l, n))
for d in order:
    by_div[d].sort(key=lambda x: (x[0], x[1]))
assert sum(len(v) for v in by_div.values()) == 46, sum(len(v) for v in by_div.values())

# Split divisions into two tables, left (Scaled) and right (RX)
LEFT_DIVS  = ["Scaled (F)", "Scaled (M)"]
RIGHT_DIVS = ["RX (F)",     "RX (M)"]

# ------------------------------------------------------------
# Render
# ------------------------------------------------------------
c = canvas.Canvas(OUT, pagesize=landscape(A3))

# page bg
c.setFillColor(PAGE_BG)
c.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)

# ============ HEADER BAND ============
HEADER_H = 0.95 * inch
y_hdr_b = PAGE_H - OUTER - HEADER_H

c.setFillColor(INK)
c.rect(OUTER, y_hdr_b, PAGE_W - 2*OUTER, HEADER_H, fill=1, stroke=0)
c.setFillColor(BANYAN)
c.rect(OUTER, y_hdr_b - 0.10 * inch, PAGE_W - 2*OUTER, 0.10 * inch,
       fill=1, stroke=0)

logo_size = 0.78 * inch
lx = M
ly = y_hdr_b + (HEADER_H - logo_size) / 2
if os.path.exists(LOGO):
    c.drawImage(LOGO, lx, ly, width=logo_size, height=logo_size,
                mask="auto", preserveAspectRatio=True)
else:
    c.setStrokeColor(BANYAN); c.setLineWidth(2)
    c.rect(lx, ly, logo_size, logo_size, stroke=1, fill=0)
    c.setFillColor(BANYAN); c.setFont("Helvetica-Bold", 24)
    c.drawCentredString(lx + logo_size/2, ly + 0.24 * inch, "BT")

wx = lx + logo_size + 0.28 * inch
c.setFillColor(white); c.setFont("Helvetica-Bold", 32)
c.drawString(wx, y_hdr_b + 0.42 * inch, "BANYAN THROWDOWN")
c.setFillColor(HexColor("#CFCFCF")); c.setFont(JP, 13)
c.drawString(wx, y_hdr_b + 0.18 * inch, "バニヤン・スローダウン")

c.setFillColor(white); c.setFont("Helvetica-Bold", 13)
c.drawRightString(PAGE_W - M, y_hdr_b + 0.62 * inch, "MASTER TALLY SHEET")
c.setFillColor(HexColor("#CFCFCF")); c.setFont(JP, 11)
c.drawRightString(PAGE_W - M, y_hdr_b + 0.42 * inch, "総合採点シート")
c.setFillColor(ACCENT); c.setFont("Helvetica-Bold", 11)
c.drawRightString(PAGE_W - M, y_hdr_b + 0.20 * inch,
                  "4/25 (Sat)  ·  @banyan_throwdown")

# ============ TABLE COLUMNS ============
cols = [
    ("H/L",         "ヒート/レーン",  0.60),
    ("ATHLETE",     "選手名",         1.95),
    ("E1",          "10-Min AMRAP",   1.00),
    ("E2·A1",       "2-Min Steps",    0.85),
    ("E2·A2",       "1RM lb",         0.85),
    ("E3",          "For Time",       0.85),
    ("E4",          "For Time 5′",    0.85),
    ("NOTES",       "メモ",           1.20),
]

# Compute per-table layout
GAP = 0.25 * inch
avail_w = (PAGE_W - 2*M - GAP) / 2  # width per table (left and right)
total_col_in = sum(w for _, _, w in cols)
scale = (avail_w / inch) / total_col_in
col_widths = [w * scale * inch for _, _, w in cols]

def col_x_positions(left):
    xs = [left]
    for w in col_widths[:-1]:
        xs.append(xs[-1] + w)
    return xs

left_l  = M
right_l = M + avail_w + GAP

# ============ DRAW A TABLE ============
HDR_ROW_H = 0.44 * inch
DIV_HDR_H = 0.32 * inch
ROW_H     = 0.33 * inch

y_table_top = y_hdr_b - 0.28 * inch  # top of the column-header row

def draw_table(x_left, divs):
    """Draw one sub-table (left or right) and return the bottom-y reached."""
    xs = col_x_positions(x_left)
    table_w = avail_w
    # column-header row
    y_ch_b = y_table_top - HDR_ROW_H
    c.setFillColor(BANYAN_D)
    c.rect(x_left, y_ch_b, table_w, HDR_ROW_H, fill=1, stroke=0)
    for i, (en, jp, _) in enumerate(cols):
        cx = xs[i] + 0.06 * inch
        c.setFillColor(white); c.setFont("Helvetica-Bold", 9.5)
        c.drawString(cx, y_ch_b + 0.25 * inch, en)
        c.setFillColor(HexColor("#B8D4B8")); c.setFont(JP, 7.5)
        c.drawString(cx, y_ch_b + 0.10 * inch, jp)
    c.setStrokeColor(white); c.setLineWidth(0.5)
    for i in range(1, len(cols)):
        c.line(xs[i], y_ch_b, xs[i], y_ch_b + HDR_ROW_H)

    y = y_ch_b
    DIV_JP = {
        "Scaled (F)": "スケールド 女子",
        "Scaled (M)": "スケールド 男子",
        "RX (F)":     "RX 女子",
        "RX (M)":     "RX 男子",
    }
    for div in divs:
        athletes = by_div[div]
        if not athletes: continue
        y -= DIV_HDR_H
        c.setFillColor(BANYAN)
        c.rect(x_left, y, table_w, DIV_HDR_H, fill=1, stroke=0)
        c.setFillColor(white); c.setFont("Helvetica-Bold", 11)
        c.drawString(x_left + 0.10 * inch, y + 0.11 * inch,
                     f"{div.upper()}   ·   {len(athletes)} athletes")
        c.setFillColor(BANYAN_L); c.setFont(JP, 9)
        c.drawRightString(x_left + table_w - 0.10 * inch, y + 0.11 * inch,
                          f"{DIV_JP[div]}  ({len(athletes)}名)")
        # athlete rows
        for idx, (h, l, name) in enumerate(athletes):
            y -= ROW_H
            if idx % 2 == 1:
                c.setFillColor(ROW_ALT)
                c.rect(x_left, y, table_w, ROW_H, fill=1, stroke=0)
            c.setFillColor(INK); c.setFont("Helvetica-Bold", 9)
            c.drawString(xs[0] + 0.08 * inch, y + 0.12 * inch, f"H{h} L{l}")
            c.setFont("Helvetica", 11)
            # truncate name to column width
            max_w = col_widths[1] - 0.16 * inch
            fs = 11
            while fs > 8 and c.stringWidth(name, "Helvetica", fs) > max_w:
                fs -= 0.5
            c.setFont("Helvetica", fs)
            c.drawString(xs[1] + 0.08 * inch, y + 0.12 * inch, name)
            # row baseline rule
            c.setStrokeColor(RULE); c.setLineWidth(0.3)
            c.line(x_left, y, x_left + table_w, y)
            # vertical grid
            c.setStrokeColor(GRID); c.setLineWidth(0.25)
            for i in range(1, len(cols)):
                c.line(xs[i], y, xs[i], y + ROW_H)

    # outer frame
    c.setStrokeColor(INK); c.setLineWidth(1.3)
    c.rect(x_left, y, table_w, y_table_top - y, stroke=1, fill=0)
    # hairline under column header
    c.line(x_left, y_ch_b, x_left + table_w, y_ch_b)
    return y

y_left  = draw_table(left_l,  LEFT_DIVS)
y_right = draw_table(right_l, RIGHT_DIVS)
y_bot = min(y_left, y_right)

# ============ FOOTER ============
foot_y = OUTER + 0.08 * inch
c.setFillColor(MUTED); c.setFont("Helvetica", 9)
c.drawString(M, foot_y + 0.16 * inch,
             "Fill in each athlete's result as the scorecard arrives.  "
             "Rank per division after each event.")
c.setFont(JP, 9)
c.drawString(M, foot_y,
             "選手の採点カードが届いたら記入してください。各イベント終了後に部門別で順位を計算。")
c.setFillColor(INK); c.setFont("Helvetica-Bold", 10)
c.drawRightString(PAGE_W - M, foot_y + 0.08 * inch,
                  "BANYAN THROWDOWN  ·  OFFICIAL TALLY")

# outer page frame
c.setStrokeColor(INK); c.setLineWidth(2.8)
c.rect(OUTER, OUTER, PAGE_W - 2*OUTER, PAGE_H - 2*OUTER, stroke=1, fill=0)

c.showPage()
c.save()
print("wrote", OUT)
