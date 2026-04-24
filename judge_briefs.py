"""
Banyan Throwdown — Individual Judge Briefs.
One PDF + one high-res PNG per judge, sized A4 portrait for
clean mobile viewing (save to photos, pinch-zoom, no scrolling).

Cross-references the judge Schedule with the athlete Schedule
so each row lists: Time · Event · Heat · Lane · Division · Athlete.
"""
import glob
import os
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor, white
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

HERE = os.path.dirname(os.path.abspath(__file__))
LOGO = os.path.join(HERE, "logo.png")

pdfmetrics.registerFont(
    TTFont("IPAGothic", "/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf")
)
JP = "IPAGothic"

INK      = HexColor("#0E0E0E")
MUTED    = HexColor("#555555")
BANYAN   = HexColor("#2E6B34")
BANYAN_D = HexColor("#1F4D2B")
BANYAN_L = HexColor("#E6F0E5")
ACCENT   = HexColor("#E07A1F")
PAGE_BG  = HexColor("#ECE8DD")
ROW_ALT  = HexColor("#F5F2E8")
RULE     = HexColor("#8A8272")
GRID     = HexColor("#D0C9B8")

OUTER = 0.18 * inch
M     = 0.32 * inch

EVENT_ROWS = {1: range(6, 15), 2: range(16, 25),
              3: range(28, 37), 4: range(38, 47)}
EVENT_TITLE = {
    1: ("EVENT 1  ·  10-Minute AMRAP",           "10分間AMRAP"),
    2: ("EVENT 2  ·  2-Min AMRAP + 1RM C&J",     "2分AMRAP + 1RM ハングC&J"),
    3: ("EVENT 3  ·  For Time (Burpee + Row)",   "タイム（バーピー＋ローワー）"),
    4: ("EVENT 4  ·  For Time (HSPU/HRPU + DL)", "タイム（HSPU/HRPU＋DL）"),
}

def load_workbook():
    path = glob.glob(os.path.join(HERE, "*.xlsx"))[0]
    return openpyxl.load_workbook(path, data_only=True)

def load_assignments(wb, judge_name):
    """Return [(event, heat, lane, division, time, athlete)] for this judge."""
    jsh = wb["judge Schedule"]
    ash = wb["Schedule"]
    needle = judge_name.strip().lower()
    out = []
    for event, rows in EVENT_ROWS.items():
        for r in rows:
            t   = jsh.cell(row=r, column=2).value
            div = jsh.cell(row=r, column=3).value
            heat = jsh.cell(row=r, column=4).value
            heat_num = int(heat.split()[1])
            for lane_idx in range(1, 7):
                j = jsh.cell(row=r, column=4 + lane_idx).value
                if j and j.strip().lower() == needle:
                    athlete = ash.cell(row=r, column=4 + lane_idx).value
                    out.append((event, heat_num, lane_idx, div,
                                t.strftime("%H:%M") if t else "", athlete or ""))
    return out

def render_brief(judge_name, judge_jp, assignments, out_pdf):
    c = canvas.Canvas(out_pdf, pagesize=A4)
    PAGE_W, PAGE_H = A4

    # bg
    c.setFillColor(PAGE_BG)
    c.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)

    # header
    HEADER_H = 0.95 * inch
    y_hdr_b = PAGE_H - OUTER - HEADER_H
    c.setFillColor(INK)
    c.rect(OUTER, y_hdr_b, PAGE_W - 2*OUTER, HEADER_H, fill=1, stroke=0)
    c.setFillColor(BANYAN)
    c.rect(OUTER, y_hdr_b - 0.08*inch, PAGE_W - 2*OUTER, 0.08*inch,
           fill=1, stroke=0)

    # logo
    logo_size = 0.72 * inch
    lx = M
    ly = y_hdr_b + (HEADER_H - logo_size) / 2
    if os.path.exists(LOGO):
        c.drawImage(LOGO, lx, ly, width=logo_size, height=logo_size,
                    mask="auto", preserveAspectRatio=True)
    else:
        c.setStrokeColor(BANYAN); c.setLineWidth(2)
        c.rect(lx, ly, logo_size, logo_size, stroke=1, fill=0)
        c.setFillColor(BANYAN); c.setFont("Helvetica-Bold", 22)
        c.drawCentredString(lx + logo_size/2, ly + 0.22*inch, "BT")

    wx = lx + logo_size + 0.22*inch
    c.setFillColor(white); c.setFont("Helvetica-Bold", 22)
    c.drawString(wx, y_hdr_b + 0.52*inch, "BANYAN THROWDOWN")
    c.setFillColor(HexColor("#CFCFCF")); c.setFont(JP, 10)
    c.drawString(wx, y_hdr_b + 0.32*inch, "バニヤン・スローダウン")
    c.setFillColor(ACCENT); c.setFont("Helvetica-Bold", 11)
    c.drawString(wx, y_hdr_b + 0.12*inch, "JUDGE BRIEF  ·  ジャッジ担当表")

    # judge name block (right)
    c.setFillColor(HexColor("#CFCFCF")); c.setFont("Helvetica-Bold", 9)
    c.drawRightString(PAGE_W - M, y_hdr_b + 0.66*inch, "JUDGE / ジャッジ")
    c.setFillColor(white); c.setFont("Helvetica-Bold", 20)
    c.drawRightString(PAGE_W - M, y_hdr_b + 0.40*inch, judge_name.upper())
    c.setFillColor(ACCENT); c.setFont(JP, 11)
    c.drawRightString(PAGE_W - M, y_hdr_b + 0.18*inch, judge_jp)

    # summary strip
    y_sum = y_hdr_b - 0.20 * inch
    c.setFillColor(MUTED); c.setFont("Helvetica-Bold", 9)
    c.drawString(M, y_sum,
                 f"{len(assignments)} ASSIGNMENTS  ·  4/25 (Sat)  ·  @banyan_throwdown")
    c.setFont(JP, 9)
    c.drawRightString(PAGE_W - M, y_sum, f"担当 {len(assignments)} 件")

    # table columns (fit A4 - 2M)
    table_top = y_sum - 0.18 * inch
    cols = [
        ("TIME",     "時間",    0.70),
        ("EVT",      "E",       0.40),
        ("HEAT",     "ヒート",  0.50),
        ("LANE",     "レーン",  0.50),
        ("DIVISION", "部門",    1.05),
        ("ATHLETE",  "選手",    3.55),
    ]
    total_in = sum(w for _, _, w in cols)
    avail_in = (PAGE_W - 2*M) / inch
    scale = avail_in / total_in
    col_w = [w * scale * inch for _, _, w in cols]
    col_x = [M]
    for w in col_w[:-1]:
        col_x.append(col_x[-1] + w)

    # column header
    CH_H = 0.38 * inch
    y_ch_b = table_top - CH_H
    c.setFillColor(BANYAN_D)
    c.rect(M, y_ch_b, PAGE_W - 2*M, CH_H, fill=1, stroke=0)
    for i, (en, jp, _) in enumerate(cols):
        cx = col_x[i] + 0.06 * inch
        c.setFillColor(white); c.setFont("Helvetica-Bold", 9)
        c.drawString(cx, y_ch_b + 0.22 * inch, en)
        c.setFillColor(HexColor("#B8D4B8")); c.setFont(JP, 7.5)
        c.drawString(cx, y_ch_b + 0.08 * inch, jp)
    c.setStrokeColor(white); c.setLineWidth(0.4)
    for i in range(1, len(cols)):
        c.line(col_x[i], y_ch_b, col_x[i], y_ch_b + CH_H)

    # rows — group by event (colored event divider between groups)
    y = y_ch_b
    ROW_H = 0.32 * inch
    DIV_H = 0.30 * inch
    by_event = {}
    for a in assignments:
        by_event.setdefault(a[0], []).append(a)

    for ev in sorted(by_event.keys()):
        rows_here = by_event[ev]
        # event divider
        y -= DIV_H
        c.setFillColor(BANYAN)
        c.rect(M, y, PAGE_W - 2*M, DIV_H, fill=1, stroke=0)
        en_title, jp_title = EVENT_TITLE[ev]
        c.setFillColor(white); c.setFont("Helvetica-Bold", 10)
        c.drawString(M + 0.10*inch, y + 0.10*inch, en_title)
        c.setFillColor(BANYAN_L); c.setFont(JP, 9)
        c.drawRightString(PAGE_W - M - 0.10*inch, y + 0.10*inch, jp_title)

        for idx, (_, h, l, div, t, athlete) in enumerate(rows_here):
            y -= ROW_H
            if idx % 2 == 1:
                c.setFillColor(ROW_ALT)
                c.rect(M, y, PAGE_W - 2*M, ROW_H, fill=1, stroke=0)
            c.setFillColor(INK); c.setFont("Helvetica-Bold", 11)
            c.drawString(col_x[0] + 0.06*inch, y + 0.12*inch, t)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(col_x[1] + 0.06*inch, y + 0.12*inch, str(ev))
            c.setFont("Helvetica", 10)
            c.drawString(col_x[2] + 0.06*inch, y + 0.12*inch, f"H{h}")
            c.setFont("Helvetica-Bold", 12)
            c.setFillColor(ACCENT)
            c.drawString(col_x[3] + 0.06*inch, y + 0.12*inch, f"L{l}")
            c.setFillColor(INK); c.setFont("Helvetica", 10)
            c.drawString(col_x[4] + 0.06*inch, y + 0.12*inch, div)
            c.setFont("Helvetica-Bold", 12)
            c.drawString(col_x[5] + 0.06*inch, y + 0.12*inch, athlete)
            # baseline rule
            c.setStrokeColor(RULE); c.setLineWidth(0.3)
            c.line(M, y, PAGE_W - M, y)
            # vertical grid
            c.setStrokeColor(GRID); c.setLineWidth(0.2)
            for i in range(1, len(cols)):
                c.line(col_x[i], y, col_x[i], y + ROW_H)

    # outer frame of table
    c.setStrokeColor(INK); c.setLineWidth(1.3)
    c.rect(M, y, PAGE_W - 2*M, table_top - y, stroke=1, fill=0)
    c.line(M, y_ch_b, PAGE_W - M, y_ch_b)

    # footer
    foot_y = OUTER + 0.08 * inch
    c.setFillColor(MUTED); c.setFont("Helvetica", 8)
    c.drawString(M, foot_y + 0.15*inch,
                 "Keep this on your phone during the event.  Report to your "
                 "assigned lane 2 minutes before heat start.")
    c.setFont(JP, 8)
    c.drawString(M, foot_y,
                 "大会中はこのリストを携帯で確認してください。ヒート開始2分前までに担当レーンへ。")
    c.setFont("Helvetica-Bold", 9); c.setFillColor(INK)
    c.drawRightString(PAGE_W - M, foot_y + 0.07*inch,
                      "BANYAN THROWDOWN  ·  JUDGE BRIEF")

    # outer page frame
    c.setStrokeColor(INK); c.setLineWidth(2.5)
    c.rect(OUTER, OUTER, PAGE_W - 2*OUTER, PAGE_H - 2*OUTER,
           stroke=1, fill=0)

    c.showPage()
    c.save()

# ---- run for both judges ----
wb = load_workbook()

for judge_en, judge_jp in [("Anthony", "アンソニー"), ("Takaya", "タカヤ")]:
    assigns = load_assignments(wb, judge_en)
    out_pdf = os.path.join(HERE, f"JudgeBrief_{judge_en}.pdf")
    render_brief(judge_en, judge_jp, assigns, out_pdf)
    print(f"{judge_en}: {len(assigns)} assignments -> {out_pdf}")
